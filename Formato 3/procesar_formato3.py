"""
Script para actualizar el FORMATO 3 de "El Papá de los formatos" cruzando datos
con PEPC y BOM por Odoo ID.

Reglas:
  - Filas de F3 cuyo Odoo ID no aparezca ni en PEPC (Código Odoo) ni en BOM (ID)
    se eliminan.
  - Si coincide con BOM: Cantidad <- CANTIDAD de BOM.
  - Si coincide con PEPC: Valor total en COP (Sin incluir IVA) <- PRECIO TOTAL
    (convertido desde USD a COP usando la TRM tomada de D7 cuando MONEDA == USD).
  - Valor IVA en COP = 19% del Valor total, EXCEPTO para paneles/módulos
    fotovoltaicos e inversores/microinversores donde es 0.
  - Si un Odoo ID está duplicado en PEPC o BOM, se toma el valor MÁS ALTO y se
    advierte.
  - Alerta de coincidencias parciales (ID en F3 que matchea solo con PEPC o
    solo con BOM, no con ambos).
  - Se sobreescribe la hoja FORMATO 3 conservando el encabezado y formato
    original; solo se reemplaza el contenido bajo el header.

Uso:
    python procesar_formato3.py
"""

import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from rapidfuzz import fuzz, process

# =====================================================================
# RUTAS DE ARCHIVOS — solo cambia los nombres de archivo
# =====================================================================
DIR_FORMATO3 = Path(__file__).parent  # siempre apunta a la carpeta Formato 3 del repo

PATH_PAPA   = DIR_FORMATO3 / "El_Papá_de_los_formatos_CON_ODOO.xlsx"
PATH_PEPC   = DIR_FORMATO3 / "PEPC 2.0_Magangué Oc_MGS_0054_CFM.xlsx"
PATH_BOM    = DIR_FORMATO3 / "BOM_MAGANGUÉOC1.xlsx"
PATH_SALIDA = DIR_FORMATO3 / "El_Papá_de_los_formatos_ACTUALIZADO.xlsx"

# =====================================================================
# COMPLETAR CÓDIGO ODOO (opcional)
#
# Cambia COMPLETAR_PEPC a True si tu PEPC no tiene la columna
# "Código Odoo" o la tiene incompleta. El script la completará
# antes de procesar el Formato 3.
# =====================================================================
COMPLETAR_PEPC = True

# PEPC de referencia: uno que ya tenga "Código Odoo" completo.
# Déjalo en None si no tienes uno; el matching usará solo BOM y Papá.
# Si lo necesitas, descomenta la línea siguiente y ajusta la ruta:
PATH_PEPC_REFERENCIA: Path | None = None
PATH_PEPC_REFERENCIA = DIR_FORMATO3 / "PEPC_1P_GENERICO_NOCFM.xlsx"

# El PEPC completado se guardará aquí (solo se usa si COMPLETAR_PEPC = True).
PATH_PEPC_COMPLETADO = PATH_PEPC.parent / (PATH_PEPC.stem + "_completado.xlsx")

# Umbral de similitud fuzzy para completar códigos (0–100).
# Sube este valor si ves demasiados falsos positivos.
UMBRAL_SIMILITUD = 50.0

# =====================================================================
# CONFIGURACIÓN INTERNA — usados como fallback si la auto-detección falla
# =====================================================================

# Fila donde está el encabezado en cada hoja (1-indexed, como en Excel)
# PEPC 2.0 Magangué tiene el header en fila 22; ajusta si usas otro archivo.
HEADER_ROW_PEPC = 22
HEADER_ROW_F3   = 11
HEADER_ROW_BOM  = 1

# Nombres de columnas en PEPC y BOM
COL_PEPC_ID      = "Código Odoo"
COL_PEPC_PRECIO  = "PRECIO TOTAL"
COL_PEPC_MONEDA  = "MONEDA"
COL_BOM_ID       = "ID"
COL_BOM_CANTIDAD = "CANTIDAD"

# Nombres en Formato 3 exentos de IVA (se pone 0)
NOMBRES_IVA_CERO = {
    "Paneles/modulos o celdas fotovoltaicas",
    "Inversores o microinversores (Off Gid, Grid Tie o Híbrido)",
}

# Valores que NO son IDs reales y deben ignorarse
IDS_INVALIDOS = {
    "", "0", "no encontrado", "no creado en odoo", "no se encuentra item", "nan",
}

IVA_RATE = 0.19

# Columnas clave para auto-detección de encabezados
# PEPC: "Código Odoo" puede no existir en el archivo objetivo (se crea después),
# por eso solo se exigen las columnas que siempre están presentes.
_CLAVES_PEPC_REQ = ["DESCRIPCIÓN", "MONEDA", "PRECIO TOTAL"]
_CLAVES_PEPC     = _CLAVES_PEPC_REQ   # alias para compatibilidad
_CLAVES_BOM      = ["MATERIAL", "ID", "CANTIDAD"]
_CLAVES_F3       = ["Nombre del Elemento", "Odoo ID", "Cantidad"]


# =====================================================================
# HELPERS
# =====================================================================

def encontrar_columna(df, fragmento: str) -> str:
    """Busca la columna cuyo nombre contenga el fragmento (ignora
    mayúsculas, saltos de línea y espacios)."""
    fragmento_norm = fragmento.lower().replace("\n", "").replace(" ", "")
    for col in df.columns:
        col_norm = str(col).lower().replace("\n", "").replace(" ", "")
        if fragmento_norm in col_norm:
            return col
    raise KeyError(
        f"No se encontró columna con fragmento: {fragmento!r}. "
        f"Columnas disponibles: {list(df.columns)}"
    )


def limpiar_id(valor) -> str | None:
    """Normaliza un Odoo ID. Devuelve el ID limpio en mayúsculas o None
    si el valor no es un ID válido."""
    if pd.isna(valor):
        return None
    s = str(valor).strip()
    if "," in s:          # IDs con coma no son válidos
        return None
    if s.lower() in IDS_INVALIDOS:
        return None
    s = re.sub(r"\s+", "", s)   # eliminar espacios/tabs internos
    if s.lower() in IDS_INVALIDOS:
        return None
    return s.upper()


def extraer_trm(valor_celda) -> float:
    """Extrae el número de TRM desde el valor de una celda del PEPC.
    Soporta: 'TRM: $3700', 'TRM: $3.700', 'TRM: 3,700', '3900', 4100.0, etc."""
    if valor_celda is None:
        raise ValueError("Celda de TRM vacía: no se puede determinar la TRM.")
    if isinstance(valor_celda, (int, float)):
        return float(valor_celda)
    s = str(valor_celda).replace("$", "").replace(" ", "")
    m = re.search(r"([\d][0-9.,]*)", s)
    if not m:
        raise ValueError(f"No se pudo extraer la TRM del valor: {valor_celda!r}")
    raw = m.group(1)
    raw_clean = re.sub(r"[.,](?=\d{3}(?:[.,]|$))", "", raw)
    raw_clean = raw_clean.replace(",", ".")
    return float(raw_clean)


def detectar_header(path, sheet_name, columnas_req, max_rows=50) -> int:
    """Auto-detecta la fila del encabezado en una hoja Excel.

    Escanea filas 0–max_rows buscando la primera en que aparezcan TODAS las
    columnas_req (coincidencia parcial, sin mayúsculas, sin tildes, sin espacios).

    Retorna el índice 0-based para pasar como header= en pd.read_excel.
    Lanza ValueError con muestra de lo encontrado si no halla el encabezado.
    """
    import unicodedata

    def _norm(s):
        s = str(s).lower().strip().replace("\n", "").replace(" ", "")
        return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()

    df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=max_rows)
    claves_norm = [_norm(c) for c in columnas_req]

    for i in range(len(df_raw)):
        row_vals = [_norm(v) for v in df_raw.iloc[i].dropna().values]
        if all(any(clave in val for val in row_vals) for clave in claves_norm):
            print(f"  Header detectado en fila Excel {i + 1} (índice 0-based: {i}) "
                  f"de '{sheet_name}' en '{Path(path).name}'")
            return i

    sample = []
    for i in range(len(df_raw)):
        vals = [str(v) for v in df_raw.iloc[i].dropna().values if str(v).strip()]
        if vals:
            sample.append(f"    Fila {i + 1}: {vals[:6]}")

    raise ValueError(
        f"No se encontró fila con columnas {columnas_req!r} en las primeras "
        f"{max_rows} filas de hoja '{sheet_name}' en '{Path(path).name}'.\n"
        "Primeras filas no vacías:\n" + "\n".join(sample[:10])
    )


def detectar_trm(path, sheet_name, max_rows=30) -> float:
    """Auto-detecta la TRM buscando una celda con texto 'TRM'.

    Soporta dos formatos:
    - Valor embebido: celda dice 'TRM: $3700' → extrae de ahí.
    - Valor separado: celda dice 'TRM' y la celda de abajo tiene el número.

    Retorna el valor numérico pasado por extraer_trm().
    Lanza ValueError si no encuentra ninguna etiqueta 'TRM' con valor.
    """
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=max_rows):
            for cell in row:
                if cell.value is None or "trm" not in str(cell.value).lower():
                    continue
                # Intentar extraer número de la misma celda
                try:
                    valor = extraer_trm(cell.value)
                    print(f"  TRM encontrada en {cell.coordinate} "
                          f"(valor embebido: {cell.value!r})")
                    return valor
                except ValueError:
                    pass
                # Si no tiene número, buscar en la celda de abajo
                below = ws.cell(row=cell.row + 1, column=cell.column)
                if below.value is not None:
                    try:
                        valor = extraer_trm(below.value)
                        print(f"  TRM encontrada en {cell.coordinate}, "
                              f"valor en {below.coordinate}: {below.value!r}")
                        return valor
                    except ValueError:
                        pass
    finally:
        wb.close()

    raise ValueError(
        f"No se encontró etiqueta 'TRM' con valor numérico en las primeras "
        f"{max_rows} filas de hoja '{sheet_name}' en '{Path(path).name}'."
    )


def reducir_max(df, col_id, col_valor, label):
    """Agrupa por ID tomando el valor máximo. Devuelve un dict
    {id: max_value} y lista de (id, [valores]) para IDs duplicados."""
    sub = df[[col_id, col_valor]].dropna(subset=[col_id])
    sub = sub[sub[col_valor].notna()].copy()
    sub[col_valor] = pd.to_numeric(sub[col_valor], errors="coerce")
    sub = sub.dropna(subset=[col_valor])

    duplicados = []
    grupos = sub.groupby(col_id)[col_valor]
    for id_, valores in grupos:
        vals = valores.tolist()
        if len(vals) > 1 and len(set(vals)) > 1:
            duplicados.append((id_, vals))
    return grupos.max().to_dict(), duplicados


# =====================================================================
# PROCESAR FORMATO 3
# =====================================================================

def procesar(df_pepc, df_f3, df_bom, trm):
    df_f3 = df_f3.copy()

    # Detectar columnas de Formato 3
    COL_F3_VALOR_TOT = encontrar_columna(df_f3, "Valor total en COP")
    COL_F3_IVA       = encontrar_columna(df_f3, "Valor IVA en COP")
    COL_F3_CANTIDAD  = encontrar_columna(df_f3, "Cantidad")
    COL_F3_NOMBRE    = encontrar_columna(df_f3, "Nombre del Elemento")
    COL_F3_ID        = encontrar_columna(df_f3, "Odoo ID")

    df_f3["_id_norm"]  = df_f3[COL_F3_ID].apply(limpiar_id)
    df_pepc = df_pepc.copy()
    df_pepc["_id_norm"] = df_pepc[COL_PEPC_ID].apply(limpiar_id)
    df_bom = df_bom.copy()
    df_bom["_id_norm"]  = df_bom[COL_BOM_ID].apply(limpiar_id)

    # Convertir PRECIO TOTAL a COP según MONEDA
    def convertir_a_cop(row):
        precio = pd.to_numeric(row[COL_PEPC_PRECIO], errors="coerce")
        if pd.isna(precio):
            return None
        moneda = str(row[COL_PEPC_MONEDA]).strip().upper() if pd.notna(row[COL_PEPC_MONEDA]) else ""
        return precio * trm if moneda == "USD" else precio

    df_pepc["_precio_cop"] = df_pepc.apply(convertir_a_cop, axis=1)

    ids_f3 = set(df_f3["_id_norm"].dropna().unique())

    pepc_filtrado = df_pepc[df_pepc["_id_norm"].isin(ids_f3)]
    bom_filtrado  = df_bom[df_bom["_id_norm"].isin(ids_f3)]

    precios_pepc,  dup_pepc = reducir_max(pepc_filtrado, "_id_norm", "_precio_cop",    "PEPC")
    cantidades_bom, dup_bom = reducir_max(bom_filtrado,  "_id_norm", COL_BOM_CANTIDAD, "BOM")

    ids_pepc = set(precios_pepc.keys())
    ids_bom  = set(cantidades_bom.keys())

    # Alertas de duplicados
    if dup_pepc:
        print("\n⚠ ALERTA: IDs con múltiples PRECIO TOTAL distintos en PEPC (se toma el MÁS ALTO):")
        print(pd.DataFrame(
            [(i, vals, max(vals)) for i, vals in dup_pepc],
            columns=["Odoo ID", "Valores encontrados (COP)", "Valor usado (máx)"],
        ).to_string(index=False))

    if dup_bom:
        print("\n⚠ ALERTA: IDs con múltiples CANTIDAD distintas en BOM (se toma la MÁS ALTA):")
        print(pd.DataFrame(
            [(i, vals, max(vals)) for i, vals in dup_bom],
            columns=["Odoo ID", "Cantidades encontradas", "Cantidad usada (máx)"],
        ).to_string(index=False))

    # Filtrar filas de F3 sin coincidencia en ninguna fuente
    en_alguno = df_f3["_id_norm"].isin(ids_pepc | ids_bom) & df_f3["_id_norm"].notna()
    eliminadas = df_f3[~en_alguno]
    df_f3 = df_f3[en_alguno].copy()

    if len(eliminadas):
        print(f"\nℹ Se eliminaron {len(eliminadas)} filas de FORMATO 3 sin "
              f"coincidencia de Odoo ID en PEPC ni BOM.")

    # Reemplazar Cantidad y Valor total (0 si no hay coincidencia en la fuente)
    df_f3[COL_F3_CANTIDAD]  = df_f3.apply(lambda r: cantidades_bom.get(r["_id_norm"], 0), axis=1)
    df_f3[COL_F3_VALOR_TOT] = df_f3.apply(lambda r: precios_pepc.get(r["_id_norm"], 0),   axis=1)

    # Recalcular IVA
    def calcular_iva(row):
        nombre = str(row[COL_F3_NOMBRE]).strip() if pd.notna(row[COL_F3_NOMBRE]) else ""
        if nombre in NOMBRES_IVA_CERO:
            return 0
        valor = pd.to_numeric(row[COL_F3_VALOR_TOT], errors="coerce")
        return None if pd.isna(valor) else valor * IVA_RATE

    df_f3[COL_F3_IVA] = df_f3.apply(calcular_iva, axis=1)

    # Alerta de coincidencias parciales
    ids_f3_final = set(df_f3["_id_norm"].dropna().unique())
    solo_pepc = (ids_f3_final & ids_pepc) - ids_bom
    solo_bom  = (ids_f3_final & ids_bom)  - ids_pepc

    if solo_pepc or solo_bom:
        print("\n⚠ ALERTA: Coincidencias parciales (Odoo ID encontrado solo en "
              "uno de los dos archivos de referencia):")
        filas = [(id_, "✓", "✗") for id_ in sorted(solo_pepc)] + \
                [(id_, "✗", "✓") for id_ in sorted(solo_bom)]
        print(pd.DataFrame(filas, columns=["Odoo ID", "En PEPC", "En BOM"]).to_string(index=False))
        print(f"\n  Total parciales: {len(filas)} "
              f"(solo PEPC: {len(solo_pepc)}, solo BOM: {len(solo_bom)})")
    else:
        print("\n✓ Todos los Odoo ID coinciden en ambos archivos (PEPC y BOM).")

    df_f3 = df_f3.drop(columns=["_id_norm"])
    return df_f3


# =====================================================================
# ESCRIBIR RESULTADO EN EL PAPÁ DE LOS FORMATOS
# =====================================================================

def escribir_resultado(df_f3_final):
    """Sobreescribe la hoja FORMATO 3 conservando header y formato.
    Borra las filas de datos existentes y escribe las nuevas."""
    PATH_SALIDA.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(PATH_PAPA, PATH_SALIDA)

    # Auto-detectar fila de encabezado en FORMATO 3
    try:
        h_f3_0based = detectar_header(PATH_PAPA, "FORMATO 3", _CLAVES_F3)
        header_row_f3 = h_f3_0based + 1  # convertir a 1-based para openpyxl
    except ValueError as e:
        header_row_f3 = HEADER_ROW_F3
        print(f"  ⚠ Auto-detección de header F3 falló ({e}). "
              f"Usando fallback: fila {header_row_f3}")

    wb = load_workbook(PATH_SALIDA)
    ws = wb["FORMATO 3"]

    primera_fila_datos = header_row_f3 + 1
    if ws.max_row >= primera_fila_datos:
        ws.delete_rows(primera_fila_datos, ws.max_row - primera_fila_datos + 1)

    headers_hoja = [
        ws.cell(row=header_row_f3, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    df_orden = df_f3_final.reindex(columns=headers_hoja)

    for i, (_, row) in enumerate(df_orden.iterrows()):
        excel_row = primera_fila_datos + i
        for j, val in enumerate(row, start=1):
            ws.cell(row=excel_row, column=j, value=None if pd.isna(val) else val)

    borde = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"),  bottom=Side(style="medium"),
    )
    ultima_fila = primera_fila_datos + len(df_orden) - 1
    for fila in range(header_row_f3, ultima_fila + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=fila, column=col).border = borde

    wb.save(PATH_SALIDA)
    wb.close()


# =====================================================================
# COMPLETAR CÓDIGO ODOO EN PEPC (opcional)
# =====================================================================

def _tiene_columna_odoo(df: pd.DataFrame) -> bool:
    try:
        col = encontrar_columna(df, "Código Odoo")
    except KeyError:
        return False
    return df[col].apply(limpiar_id).notna().any()


def completar_codigo_odoo(
    path_pepc_objetivo: Path,
    path_bom: Path,
    path_papa: Path,
    path_salida_pepc: Path,
    path_pepc_referencia: Path | None = None,
    umbral_similitud: float = 50.0,
) -> dict:
    """
    Completa la columna 'Código Odoo' en un PEPC fila a fila.

    - Si la fila ya tiene un Código Odoo válido → se conserva intacto.
    - Si la fila NO tiene código → se intenta asignar uno en este orden:
        1. Coincidencia exacta de DESCRIPCIÓN vs PEPC de referencia
           (solo si path_pepc_referencia está definido).
        2. Fuzzy match de DESCRIPCIÓN vs MATERIAL de BOM.
        3. Fuzzy match de DESCRIPCIÓN vs Modelo/Referencia del Papá.
        4. Fuzzy match de DESCRIPCIÓN vs DESCRIPCIÓN del PEPC de referencia
           (solo si path_pepc_referencia está definido).
    - Si ninguna fuente supera el umbral, la celda queda vacía y se reporta.

    Devuelve un dict {índice_fila_df (0-based) → código_odoo | None} con los
    códigos finales para TODAS las filas del PEPC objetivo.  Ese dict es la
    fuente de verdad que main() inyecta en el DataFrame leído del original, de
    modo que nunca se pierde ningún valor calculado por fórmulas de Excel.

    IMPORTANTE — escritura del archivo completado:
    openpyxl guarda fórmulas como strings sin valor cacheado.  Para que el
    archivo en disco también sea usable independientemente, se leen los valores
    calculados del original con data_only=True y se escriben como literales
    junto con los nuevos códigos Odoo.  Así el .xlsx completado es autónomo.
    """
    print("\n=== completar_codigo_odoo ===")
    usar_referencia = path_pepc_referencia is not None

    # Auto-detectar filas de encabezado
    print("  Detectando encabezados...")
    try:
        h_pepc_obj = detectar_header(path_pepc_objetivo, "PEPC", _CLAVES_PEPC)
    except ValueError as e:
        h_pepc_obj = HEADER_ROW_PEPC - 1
        print(f"  ⚠ Auto-detección PEPC objetivo falló ({e}). "
              f"Usando fallback: fila {HEADER_ROW_PEPC}")

    try:
        h_bom = detectar_header(path_bom, "BOM", _CLAVES_BOM)
    except ValueError as e:
        h_bom = HEADER_ROW_BOM - 1
        print(f"  ⚠ Auto-detección BOM falló ({e}). "
              f"Usando fallback: fila {HEADER_ROW_BOM}")

    try:
        h_f3 = detectar_header(path_papa, "FORMATO 3", _CLAVES_F3)
    except ValueError as e:
        h_f3 = HEADER_ROW_F3 - 1
        print(f"  ⚠ Auto-detección FORMATO 3 falló ({e}). "
              f"Usando fallback: fila {HEADER_ROW_F3}")

    df_obj = pd.read_excel(path_pepc_objetivo, sheet_name="PEPC",      header=h_pepc_obj)
    df_bom = pd.read_excel(path_bom,           sheet_name="BOM",       header=h_bom)
    df_f3  = pd.read_excel(path_papa,          sheet_name="FORMATO 3", header=h_f3)

    if usar_referencia:
        try:
            h_pepc_ref = detectar_header(path_pepc_referencia, "PEPC", _CLAVES_PEPC)
        except ValueError as e:
            h_pepc_ref = HEADER_ROW_PEPC - 1
            print(f"  ⚠ Auto-detección PEPC referencia falló ({e}). "
                  f"Usando fallback: fila {HEADER_ROW_PEPC}")
        df_ref = pd.read_excel(path_pepc_referencia, sheet_name="PEPC", header=h_pepc_ref)
    else:
        df_ref = None

    col_desc_obj  = encontrar_columna(df_obj, "DESCRIPCIÓN")
    col_bom_mat   = encontrar_columna(df_bom, "MATERIAL")
    col_bom_id    = encontrar_columna(df_bom, "ID")
    col_f3_modelo = encontrar_columna(df_f3,  "Modelo / Referencia")
    col_f3_odoo   = encontrar_columna(df_f3,  "Odoo ID")

    try:
        col_odoo_obj = encontrar_columna(df_obj, "Código Odoo")
        col_existe = True
    except KeyError:
        col_odoo_obj = "Código Odoo"
        df_obj[col_odoo_obj] = None
        col_existe = False

    ya_tenian = df_obj[col_odoo_obj].apply(limpiar_id).notna().sum()
    modo = "parcialmente completado" if (col_existe and ya_tenian > 0) else \
           "sin columna Código Odoo" if not col_existe else "con columna vacía"
    print(f"  PEPC objetivo ({modo}): {len(df_obj)} filas, {ya_tenian} ya con código")
    print(f"  BOM:                    {len(df_bom)} filas")
    print(f"  FORMATO 3:              {len(df_f3)} filas")
    if usar_referencia:
        print(f"  PEPC referencia:        {len(df_ref)} filas")
    else:
        print("  PEPC referencia:        no proporcionado (solo BOM y Papá)")

    # Construir lookups
    lookup_ref = {}
    if usar_referencia:
        col_desc_ref = encontrar_columna(df_ref, "DESCRIPCIÓN")
        col_odoo_ref = encontrar_columna(df_ref, "Código Odoo")
        for _, row in df_ref.iterrows():
            desc = str(row[col_desc_ref]).strip() if pd.notna(row[col_desc_ref]) else ""
            cod  = limpiar_id(row[col_odoo_ref])
            if desc and cod:
                lookup_ref[desc.lower()] = (desc, cod)

    lookup_bom = {}
    for _, row in df_bom.iterrows():
        mat = str(row[col_bom_mat]).strip() if pd.notna(row[col_bom_mat]) else ""
        cod = limpiar_id(row[col_bom_id])
        if mat and cod:
            lookup_bom[mat.lower()] = (mat, cod)

    lookup_f3 = {}
    for _, row in df_f3.iterrows():
        mod = str(row[col_f3_modelo]).strip() if pd.notna(row[col_f3_modelo]) else ""
        cod = limpiar_id(row[col_f3_odoo])
        if mod and cod:
            lookup_f3[mod.lower()] = (mod, cod)

    claves_ref = list(lookup_ref.keys())
    claves_bom = list(lookup_bom.keys())
    claves_f3  = list(lookup_f3.keys())

    codigos_finales = []
    alertas = []
    stats = {"conservadas": 0, "exactas": 0, "bom": 0, "f3": 0, "fuzzy_ref": 0, "sin_match": 0}

    # h_pepc_obj + 2 convierte el índice 0-based del df al número de fila Excel (1-based)
    fila_excel_base = h_pepc_obj + 2

    for idx, row in df_obj.iterrows():
        cod_actual = limpiar_id(row[col_odoo_obj])
        if cod_actual:
            codigos_finales.append(cod_actual)
            stats["conservadas"] += 1
            continue

        desc_raw  = str(row[col_desc_obj]).strip() if pd.notna(row[col_desc_obj]) else ""
        desc_norm = desc_raw.lower()

        if not desc_norm:
            codigos_finales.append(None)
            stats["sin_match"] += 1
            continue

        if usar_referencia and desc_norm in lookup_ref:
            _, cod = lookup_ref[desc_norm]
            codigos_finales.append(cod)
            stats["exactas"] += 1
            continue

        codigo_asignado = None
        score_max = 0
        fuente_match = ""

        if claves_bom:
            m = process.extractOne(desc_norm, claves_bom, scorer=fuzz.token_set_ratio,
                                   score_cutoff=umbral_similitud)
            if m and m[1] > score_max:
                score_max = m[1]
                _, codigo_asignado = lookup_bom[m[0]]
                fuente_match = f"BOM/MATERIAL (score={m[1]:.0f}%)"

        if claves_f3:
            m = process.extractOne(desc_norm, claves_f3, scorer=fuzz.token_set_ratio,
                                   score_cutoff=umbral_similitud)
            if m and m[1] > score_max:
                score_max = m[1]
                _, codigo_asignado = lookup_f3[m[0]]
                fuente_match = f"FORMATO3/Modelo (score={m[1]:.0f}%)"

        if usar_referencia and claves_ref:
            m = process.extractOne(desc_norm, claves_ref, scorer=fuzz.token_set_ratio,
                                   score_cutoff=umbral_similitud)
            if m and m[1] > score_max:
                score_max = m[1]
                _, codigo_asignado = lookup_ref[m[0]]
                fuente_match = f"PEPC-ref/DESCRIPCIÓN-fuzzy (score={m[1]:.0f}%)"

        if codigo_asignado:
            codigos_finales.append(codigo_asignado)
            key = fuente_match.split("/")[0].lower()
            if "bom" in key:       stats["bom"] += 1
            elif "formato" in key: stats["f3"] += 1
            else:                  stats["fuzzy_ref"] += 1
        else:
            codigos_finales.append(None)
            stats["sin_match"] += 1
            alertas.append({
                "Fila Excel":  idx + fila_excel_base,
                "DESCRIPCIÓN": desc_raw,
                "Mejor score": f"{score_max:.0f}% (umbral: {umbral_similitud:.0f}%)",
            })

    print(f"\n  Resumen:")
    print(f"    Conservadas (ya tenían código):    {stats['conservadas']}")
    print(f"    Asignadas por coincidencia exacta: {stats['exactas']}")
    print(f"    Asignadas por BOM/MATERIAL:        {stats['bom']}")
    print(f"    Asignadas por FORMATO3/Modelo:     {stats['f3']}")
    print(f"    Asignadas por fuzzy PEPC-ref:      {stats['fuzzy_ref']}")
    print(f"    Sin match (quedan vacías):         {stats['sin_match']}")

    if alertas:
        print(f"\n⚠ ALERTA: {len(alertas)} filas sin Código Odoo asignado:")
        print(pd.DataFrame(alertas).to_string(index=False))
    else:
        print("\n✓ Todas las filas sin código previo obtuvieron un Código Odoo.")

    # ------------------------------------------------------------------
    # Guardar el PEPC completado en disco
    #
    # Problema: openpyxl no evalúa fórmulas. Si hacemos shutil.copy()
    # y luego load_workbook() sin data_only, las fórmulas se copian como
    # strings y al releer con pandas sus celdas devuelven NaN.
    #
    # Solución robusta:
    #   1. Leer el original con data_only=True → capturar valores ya
    #      calculados por Excel para todas las celdas de datos.
    #   2. Abrir el original sin data_only (preserva estructura/formato).
    #   3. Reemplazar cada celda de datos con su valor literal —
    #      convierte fórmulas → valores concretos.
    #   4. Escribir la columna Código Odoo con los códigos asignados.
    #   5. Guardar → el .xlsx resultante es autónomo, sin fórmulas rotas.
    # ------------------------------------------------------------------
    path_salida_pepc.parent.mkdir(parents=True, exist_ok=True)

    header_row    = h_pepc_obj + 1   # 1-based para openpyxl
    primera_datos = header_row + 1

    # Paso 1: leer valores calculados del original
    wb_vals = load_workbook(path_pepc_objetivo, data_only=True)
    ws_vals = wb_vals["PEPC"]
    max_col_orig = ws_vals.max_column
    max_row_orig = ws_vals.max_row
    valores_calculados = [
        [ws_vals.cell(row=r, column=c).value for c in range(1, max_col_orig + 1)]
        for r in range(primera_datos, max_row_orig + 1)
    ]
    wb_vals.close()

    # Paso 2: abrir original preservando formato/estructura
    shutil.copy(path_pepc_objetivo, path_salida_pepc)
    wb = load_workbook(path_salida_pepc)
    ws = wb["PEPC"]

    # Paso 3: sobreescribir datos con valores literales (elimina fórmulas)
    for r_offset, fila_vals in enumerate(valores_calculados):
        excel_row = primera_datos + r_offset
        for c_offset, val in enumerate(fila_vals):
            ws.cell(row=excel_row, column=c_offset + 1, value=val)

    # Paso 4: localizar o crear la columna Código Odoo
    col_insertar = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and "código odoo" in str(val).lower():
            col_insertar = c
            break
    if col_insertar is None:
        col_insertar = ws.max_column + 1
        ws.cell(row=header_row, column=col_insertar, value="Código Odoo")

    for i, cod in enumerate(codigos_finales):
        ws.cell(row=primera_datos + i, column=col_insertar, value=cod)

    # Paso 5: guardar
    wb.save(path_salida_pepc)
    wb.close()
    print(f"\n✓ PEPC completado guardado en: {path_salida_pepc}")
    print(  "  (fórmulas convertidas a valores literales — precios preservados)")

    # Devolver dict {índice_df (0-based) → código} para que main() lo
    # inyecte directamente en el DataFrame leído del original, sin
    # necesidad de releer el archivo completado.
    return {i: cod for i, cod in enumerate(codigos_finales)}


# =====================================================================
# MAIN
# =====================================================================

def main():
    # ── Paso 1: auto-detectar encabezados ─────────────────────────────
    print("\nDetectando encabezados...")
    try:
        h_pepc = detectar_header(PATH_PEPC, "PEPC", _CLAVES_PEPC)
    except ValueError as e:
        h_pepc = HEADER_ROW_PEPC - 1
        print(f"  ⚠ Auto-detección PEPC falló ({e}). Usando fallback: fila {HEADER_ROW_PEPC}")

    try:
        h_f3 = detectar_header(PATH_PAPA, "FORMATO 3", _CLAVES_F3)
    except ValueError as e:
        h_f3 = HEADER_ROW_F3 - 1
        print(f"  ⚠ Auto-detección FORMATO 3 falló ({e}). Usando fallback: fila {HEADER_ROW_F3}")

    try:
        h_bom = detectar_header(PATH_BOM, "BOM", _CLAVES_BOM)
    except ValueError as e:
        h_bom = HEADER_ROW_BOM - 1
        print(f"  ⚠ Auto-detección BOM falló ({e}). Usando fallback: fila {HEADER_ROW_BOM}")

    # ── Paso 2: cargar archivos ────────────────────────────────────────
    # El PEPC siempre se lee del archivo ORIGINAL para preservar los
    # valores calculados por Excel (fórmulas de precios, totales, etc.).
    # Los Códigos Odoo se inyectan en memoria después del matching,
    # sin pasar por un archivo intermedio que pierda esos valores.
    print("\nCargando archivos...")
    df_pepc = pd.read_excel(PATH_PEPC, sheet_name="PEPC",      header=h_pepc)
    df_f3   = pd.read_excel(PATH_PAPA, sheet_name="FORMATO 3", header=h_f3)
    df_bom  = pd.read_excel(PATH_BOM,  sheet_name="BOM",       header=h_bom)
    print(f"  PEPC:      {df_pepc.shape[0]} filas")
    print(f"  FORMATO 3: {df_f3.shape[0]} filas")
    print(f"  BOM:       {df_bom.shape[0]} filas")

    # ── Paso 3 (opcional): completar Código Odoo ──────────────────────
    # completar_codigo_odoo() devuelve {índice_df → código} y además
    # guarda el .xlsx completado con valores literales (sin fórmulas
    # rotas) como referencia externa.  En main() usamos el dict directo
    # para nunca depender del archivo intermedio.
    if COMPLETAR_PEPC:
        codigos_dict = completar_codigo_odoo(
            path_pepc_objetivo   = PATH_PEPC,
            path_bom             = PATH_BOM,
            path_papa            = PATH_PAPA,
            path_salida_pepc     = PATH_PEPC_COMPLETADO,
            path_pepc_referencia = PATH_PEPC_REFERENCIA,
            umbral_similitud     = UMBRAL_SIMILITUD,
        )
        # Inyectar códigos en el df leído del original
        col_odoo = COL_PEPC_ID
        if col_odoo not in df_pepc.columns:
            df_pepc[col_odoo] = None
        for idx, cod in codigos_dict.items():
            if idx < len(df_pepc):
                df_pepc.at[df_pepc.index[idx], col_odoo] = cod
        print(f"\nCódigos Odoo inyectados en DataFrame del PEPC original.")

    # ── Paso 4: obtener TRM ────────────────────────────────────────────
    print("\nDetectando TRM...")
    try:
        trm = detectar_trm(PATH_PEPC, "PEPC")
    except ValueError as e:
        print(f"  ⚠ Auto-detección de TRM falló ({e}). Usando fallback D7...")
        wb_trm = load_workbook(PATH_PEPC, data_only=True)
        try:
            trm = extraer_trm(wb_trm["PEPC"]["D7"].value)
            print(f"  TRM obtenida de D7 (fallback): {trm:,.2f}")
        finally:
            wb_trm.close()
    print(f"  TRM utilizada: {trm:,.2f}")

    # ── Paso 5: procesar Formato 3 ────────────────────────────────────
    df_final = procesar(df_pepc, df_f3, df_bom, trm)
    print(f"\nFORMATO 3 final: {df_final.shape[0]} filas")

    # ── Paso 6: exportar ──────────────────────────────────────────────
    escribir_resultado(df_final)
    print(f"\n✓ Archivo guardado en: {PATH_SALIDA}")


if __name__ == "__main__":
    main()
