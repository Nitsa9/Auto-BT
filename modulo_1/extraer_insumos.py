"""
=============================================================
  EXTRACTOR DE INSUMOS - FORMULARIO UPME
=============================================================
Consolida todos los datos necesarios para diligenciar el
formulario UPME de un proyecto solar específico, extrayendo
información de 3 fuentes:

  1. proyectos_solenium.json  → datos del proyecto (API)
  2. PDFs de ingeniería       → potencia, área, energía, PR
  3. PDFs y Excel de BT       → servicios con alcance, valor, proveedor

REQUISITOS:
    pip install pymupdf openpyxl

CÓMO USAR:
    python extraer_insumos.py
    → Genera: data/insumos_MGS_0025.json
=============================================================
"""

import json
import logging
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from difflib import SequenceMatcher
from pathlib import Path

import fitz          # PyMuPDF — pip install pymupdf
import openpyxl

# =============================================================
#  LOGGING
# =============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# =============================================================
#  RUTAS  ← ajusta BASE_DIR si tu carpeta raíz es diferente
# =============================================================
BASE_DIR = Path(__file__).parent.parent  # Auto-BT/

NOMBRE_PROYECTO   = "MGS_0025_EVA2_El Copey Occidente"
JSON_SOLENIUM     = BASE_DIR / "data" / "proyectos_solenium.json"
DIR_PROYECTO      = BASE_DIR / "data" / "MGS_0025"

PDF_LOCALIZACION  = DIR_PROYECTO / "03_Engineering/05_Final version/03_Layouts 2D" / \
                    "SDS4_Cope-CIV-PL-01_Plano de localizaciones y accesos.pdf"
PDF_INFORME_ELE   = DIR_PROYECTO / "03_Engineering/05_Final version/01_Simulation" / \
                    "Cope-INF-ELE-V2.pdf"
XLSX_BT           = DIR_PROYECTO / "06_Financial/06_BT" / "Copey definitivo.xlsx"
DIR_SERVICIOS     = DIR_PROYECTO / "06_Financial/06_BT/Servicios Definitivos"


# =============================================================
#  HELPERS
# =============================================================

def leer_pdf_texto(ruta: Path) -> str:
    """Extrae todo el texto de un PDF usando PyMuPDF."""
    try:
        doc = fitz.open(str(ruta))
        texto = "\n".join(page.get_text() for page in doc)
        doc.close()
        return texto
    except Exception as e:
        logger.error("No se pudo leer PDF %s: %s", ruta.name, e)
        return ""


def buscar_numero(patron: str, texto: str, flags=re.IGNORECASE) -> float | None:
    """
    Busca un patrón en el texto y retorna el primer número encontrado
    después de él. Soporta formatos: 1.234,56 / 1,234.56 / 1234.56
    """
    match = re.search(patron, texto, flags)
    if not match:
        return None
    # Busca el número que sigue al match
    resto = texto[match.end():]
    num_match = re.search(r"[\d][0-9.,\s]*", resto)
    if not num_match:
        return None
    raw = num_match.group().strip().replace(" ", "")
    # Normalizar separadores: detecta si la coma es decimal o de miles
    if re.search(r"\d{1,3}(\.\d{3})+(,\d+)?$", raw):
        raw = raw.replace(".", "").replace(",", ".")
    elif re.search(r"\d{1,3}(,\d{3})+(\.\d+)?$", raw):
        raw = raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def extraer_mgs_id(nombre: str) -> str:
    """
    Del nombre 'MGS_0025_EVA2_El Copey Occidente' extrae 'MGS_0025'.
    Toma 'MGS' + el primer grupo de 4 dígitos que le sigue.
    """
    match = re.search(r"(MGS[_\-]?\d{4})", nombre, re.IGNORECASE)
    return match.group(1).upper() if match else nombre


# =============================================================
#  FUENTE 1: proyectos_solenium.json
# =============================================================

def cargar_datos_solenium(nombre_proyecto: str) -> dict:
    """
    Carga el JSON de Solenium y localiza el proyecto por nombre exacto.
    Retorna el diccionario del proyecto o {} si no se encuentra.
    """
    try:
        with open(JSON_SOLENIUM, encoding="utf-8") as f:
            datos = json.load(f)
        proyectos = datos.get("proyectos", [])
        for p in proyectos:
            if (p.get("nombre") == nombre_proyecto or
                    p.get("despliegue_interno") == nombre_proyecto or
                    p.get("cuenta_analitica") == nombre_proyecto.split("_")[1] if "_" in nombre_proyecto else False):
                logger.info("Proyecto encontrado en Solenium: %s", nombre_proyecto)
                return p
        logger.warning("Proyecto '%s' no encontrado en Solenium.", nombre_proyecto)
        return {}
    except Exception as e:
        logger.error("Error leyendo proyectos_solenium.json: %s", e)
        return {}


def calcular_inicio_construccion(fpo_unergy: str | None) -> str | None:
    """
    Resta 6 meses a fpo_unergy (formato YYYY-MM-DD) para obtener
    la fecha estimada de inicio de construcción.
    """
    if not fpo_unergy:
        return None
    try:
        fecha = datetime.strptime(fpo_unergy, "%Y-%m-%d")
        inicio = fecha - relativedelta(months=6)
        return inicio.strftime("%Y-%m-%d")
    except ValueError:
        logger.warning("Formato de fecha inesperado en fpo_unergy: %s", fpo_unergy)
        return None


# =============================================================
#  FUENTE 2: PDF — Plano de localizaciones
# =============================================================

def extraer_pdf_localizacion(ruta: Path) -> dict:
    """
    Extrae del plano de localizaciones:
      - Potencia nominal (kWp)  → Capacidad instalada
      - Área cerramiento (m²)   → Área del proyecto
    """
    resultado = {"capacidad_instalada_kwp_plano": None, "area_proyecto_m2": None}
    texto = leer_pdf_texto(ruta)
    if not texto:
        return resultado

    # Potencia nominal
    potencia = buscar_numero(r"potencia\s+nominal", texto)
    if potencia is not None:
        resultado["capacidad_instalada_kwp_plano"] = potencia
        logger.info("Potencia nominal (plano): %s kWp", potencia)
    else:
        logger.warning("No se encontró 'Potencia nominal' en %s", ruta.name)

    # Área cerramiento
    area = buscar_numero(r"[áa]rea\s+cerramiento", texto)
    if area is not None:
        resultado["area_proyecto_m2"] = area
        logger.info("Área cerramiento: %s m²", area)
    else:
        logger.warning("No se encontró 'Área Cerramiento' en %s", ruta.name)

    return resultado


# =============================================================
#  FUENTE 2: PDF — Informe eléctrico (simulación)
# =============================================================

def extraer_pdf_informe_ele(ruta: Path) -> dict:
    """
    Extrae del informe eléctrico de simulación:
      - Produced Energy P99 (kWh/año) → Energía generada
      - Performance Ratio PR (%)       → Eficiencia de la planta
      - Total power (kWp)              → Capacidad instalada
    """
    resultado = {
        "energia_generada_kwh_anio": None,
        "eficiencia_planta_pct":     None,
        "capacidad_instalada_kwp":   None,
    }
    texto = leer_pdf_texto(ruta)
    if not texto:
        return resultado

    # --- Energía generada: Produced Energy (P99) ---
    # Puede venir en MWh o kWh; busca la unidad para convertir si es necesario
    match_energia = re.search(
        r"produced\s+energy.*?p99[^\d]*?([\d][0-9.,\s]*)\s*(MWh|kWh)",
        texto, re.IGNORECASE | re.DOTALL
    )
    if match_energia:
        raw_bloque = match_energia.group(1).strip()
        unidad     = match_energia.group(2).upper()
        # Puede venir como "2736.3\n2402.0\n2129.6" — P99 es el último (más conservador)
        numeros = re.findall(r"[\d]+[.,]?[\d]*", raw_bloque)
        if numeros:
            raw = numeros[-1].replace(",", ".")
            try:
                valor = float(raw)
                if unidad == "MWH":
                    valor = valor * 1000
                resultado["energia_generada_kwh_anio"] = round(valor, 2)
                logger.info("Energía generada P99: %s kWh/año", resultado["energia_generada_kwh_anio"])
            except ValueError:
                logger.warning("No se pudo convertir energía P99: %s", raw)
        else:
            logger.warning("No se pudo extraer número de energía P99: %s", raw_bloque)
    else:
        logger.warning("No se encontró 'Produced Energy P99' en %s", ruta.name)

    # --- Performance Ratio ---
    # Texto real: "Perf. Ratio PR Bifacial perf. ratio 82.35 78.54 % %"
    # El valor buscado (P50) viene justo antes del primer "%"
    match_pr = re.search(
        r"perf[\.\s]*ratio\s+pr\b.{0,60}?([\d]{2,3}(?:[.,]\d+)?)\s*%",
        texto, re.IGNORECASE | re.DOTALL
    )
    if match_pr:
        pr = float(match_pr.group(1).replace(",", "."))
        resultado["eficiencia_planta_pct"] = pr
        logger.info("Performance Ratio: %s%%", pr)
    else:
        pr = buscar_numero(r"performance\s+ratio", texto)
        if pr is not None:
            resultado["eficiencia_planta_pct"] = pr
            logger.info("Performance Ratio: %s%%", pr)
        else:
            logger.warning("No se encontró 'Perf. Ratio PR' en %s", ruta.name)

    # --- Total power ---
    # Texto real: "Total power Pnom ratio 5 990 1.33"
    # El valor de potencia es el número grande después de "ratio N"
    match_tp = re.search(
        r"total\s+power\s+pnom\s+ratio\s+\d+\s+([\d.,]+)",
        texto, re.IGNORECASE
    )
    if match_tp:
        total_power = float(match_tp.group(1).replace(",", "."))
        resultado["capacidad_instalada_kwp"] = total_power
        logger.info("Total power: %s kWp", total_power)
    else:
        total_power = buscar_numero(r"total\s+power", texto)
        if total_power is not None:
            resultado["capacidad_instalada_kwp"] = total_power
            logger.info("Total power: %s kWp", total_power)
        else:
            logger.warning("No se encontró 'Total power' en %s", ruta.name)

    return resultado


# =============================================================
#  FUENTE 3: PDFs de Servicios Definitivos + Excel BT
# =============================================================

def _parsear_valor(raw: str) -> float | None:
    """Convierte un string de valor monetario a float."""
    raw = re.sub(r"[^\d.,]", "", raw.strip())
    if not raw:
        return None
    if re.search(r"\d{1,3}(\.\d{3})+(,\d+)?$", raw):
        raw = raw.replace(".", "").replace(",", ".")
    elif re.search(r"\d{1,3}(,\d{3})+(\.\d+)?$", raw):
        raw = raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def _similitud(a: str, b: str) -> float:
    """Retorna ratio de similitud entre dos strings (0-1)."""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _extraer_filas_tabla_pdf(pdf_path: Path) -> list[dict]:
    """
    Extrae filas de servicios de un PDF usando detección de tablas espaciales.
    Identifica columnas ALCANCE y VALOR TOTAL por encabezado.
    """
    filas = []
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        logger.error("No se pudo abrir %s: %s", pdf_path.name, e)
        return filas

    for page in doc:
        try:
            tables = page.find_tables()
        except AttributeError:
            # PyMuPDF < 1.23 no tiene find_tables — fallback a texto plano
            doc.close()
            return []

        for table in tables:
            datos = table.extract()
            if not datos:
                continue

            # Busca fila de encabezados
            header_idx = None
            for i, fila in enumerate(datos[:6]):
                fila_norm = [str(c).lower().strip() if c else "" for c in fila]
                if any("alcance" in c for c in fila_norm):
                    header_idx = i
                    break
            if header_idx is None:
                continue

            headers = [str(c).lower().strip() if c else "" for c in datos[header_idx]]
            col_alc = next((j for j, h in enumerate(headers) if "alcance" in h), None)
            col_val = next((j for j, h in enumerate(headers)
                            if "valor" in h and "total" in h), None)
            col_iva = next((j for j, h in enumerate(headers) if "iva" in h), None)

            if col_alc is None:
                continue

            for fila in datos[header_idx + 1:]:
                if not any(fila):
                    continue
                alcance = str(fila[col_alc]).strip() if fila[col_alc] else ""
                if not alcance or alcance.lower() == "none":
                    continue
                # Limpia saltos de línea internos del alcance
                alcance = re.sub(r"\s*\n\s*", " ", alcance).strip()

                valor = None
                if col_val is not None and fila[col_val]:
                    valor = _parsear_valor(str(fila[col_val]))

                iva = None
                if col_iva is not None and fila[col_iva]:
                    iva = _parsear_valor(str(fila[col_iva]))
                if iva is None and valor is not None:
                    iva = round(valor * 0.19, 2)

                filas.append({
                    "alcance":         alcance,
                    "valor_total_cop": valor,
                    "valor_iva_cop":   iva,
                    "proveedor":       None,
                    "servicio":        None,
                    "_fuente_pdf":     pdf_path.name,
                })

    doc.close()
    return filas


def _extraer_filas_texto_pdf(pdf_path: Path) -> list[dict]:
    """
    Fallback: extrae servicios de un PDF por texto plano (sin tablas).
    Busca bloques ALCANCE: ... y el valor $ que les sigue.
    """
    filas = []
    texto = leer_pdf_texto(pdf_path)
    if not texto:
        return filas

    secciones = re.split(r"ALCANCE\s*:", texto, flags=re.IGNORECASE)
    bloques = secciones[1:] if len(secciones) > 1 else [texto]

    for bloque in bloques:
        alcance_match = re.match(r"(.+?)(?:\n\n|\Z)", bloque, re.DOTALL)
        alcance = alcance_match.group(1).strip() if alcance_match else bloque[:300].strip()
        alcance = re.sub(r"\s*\n\s*", " ", alcance).strip()

        valor = None
        for patron in [
            r"(?:valor\s+total|total\s+sin\s+iva|subtotal)[^\d$]*\$?\s*([\d][0-9.,\s]*)",
            r"\$\s*([\d][0-9.,\s]{4,})",
        ]:
            m = re.search(patron, bloque, re.IGNORECASE)
            if m:
                valor = _parsear_valor(m.group(1))
                if valor:
                    break

        iva = None
        m_iva = re.search(r"(?:iva|impuesto)[^\d$]*\$?\s*([\d][0-9.,\s]*)", bloque, re.IGNORECASE)
        if m_iva:
            iva = _parsear_valor(m_iva.group(1))
        if iva is None and valor is not None:
            iva = round(valor * 0.19, 2)
            logger.info("IVA calculado (19%%) para alcance de %s", pdf_path.name)

        filas.append({
            "alcance":         alcance,
            "valor_total_cop": valor,
            "valor_iva_cop":   iva,
            "proveedor":       None,
            "servicio":        None,
            "_fuente_pdf":     pdf_path.name,
        })

    return filas


def extraer_servicios_pdf(dir_servicios: Path) -> list[dict]:
    """
    Recorre todos los PDFs en Servicios Definitivos.
    Intenta extracción por tabla espacial; si falla, usa texto plano.
    """
    servicios = []
    pdfs = sorted(dir_servicios.glob("*.pdf"))

    if not pdfs:
        logger.warning("No se encontraron PDFs en %s", dir_servicios)
        return servicios

    for pdf_path in pdfs:
        logger.info("Procesando servicios PDF: %s", pdf_path.name)
        filas = _extraer_filas_tabla_pdf(pdf_path)
        if not filas:
            filas = _extraer_filas_texto_pdf(pdf_path)
        servicios.extend(filas)

    logger.info("Total de alcances extraídos de PDFs: %d", len(servicios))
    return servicios


def enriquecer_servicios_excel(servicios: list[dict], xlsx_path: Path) -> list[dict]:
    """
    Lee la hoja FORMATO 4 del Excel y enriquece cada servicio con
    Proveedor, Servicio y (si falta) Valor Total.
    Usa fuzzy matching con umbral de similitud del 30%.
    """
    UMBRAL_SIMILITUD = 0.30

    if not xlsx_path.exists():
        logger.warning("No se encontró el Excel BT: %s", xlsx_path)
        return servicios

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception as e:
        logger.error("No se pudo abrir %s: %s", xlsx_path.name, e)
        return servicios

    # Busca la hoja FORMATO 4 (nombre exacto o parcial)
    hoja = next(
        (wb[n] for n in wb.sheetnames if "formato" in n.lower() and "4" in n),
        None
    )
    if hoja is None:
        logger.warning("No se encontró hoja 'FORMATO 4' en %s", xlsx_path.name)
        return servicios

    filas_excel = list(hoja.iter_rows(values_only=True))

    # Localiza fila de encabezados
    header_idx = None
    cols = {}
    for i, fila in enumerate(filas_excel[:15]):
        fila_norm = [str(c).lower().strip() if c else "" for c in fila]
        if any("alcance" in c for c in fila_norm):
            header_idx = i
            for j, c in enumerate(fila_norm):
                if "alcance"   in c: cols["alcance"]   = j
                if "proveedor" in c: cols["proveedor"] = j
                if "servicio"  in c: cols["servicio"]  = j
                if "valor" in c and "total" in c: cols["valor"] = j
            break

    if header_idx is None or "alcance" not in cols:
        logger.warning("No se encontraron encabezados en hoja FORMATO 4")
        return servicios

    # Construye índice de filas del Excel
    filas_idx = []
    for fila in filas_excel[header_idx + 1:]:
        alc = fila[cols["alcance"]] if cols.get("alcance") is not None else None
        if not alc:
            continue
        filas_idx.append({
            "alcance":   str(alc).strip(),
            "proveedor": str(fila[cols["proveedor"]]).strip() if cols.get("proveedor") is not None and fila[cols["proveedor"]] else None,
            "servicio":  str(fila[cols["servicio"]]).strip()  if cols.get("servicio")  is not None and fila[cols["servicio"]]  else None,
            "valor":     _parsear_valor(str(fila[cols["valor"]])) if cols.get("valor") is not None and fila[cols.get("valor")] else None,
        })

    # Cruza cada servicio del PDF con el Excel por similitud
    for svc in servicios:
        mejor_ratio = 0.0
        mejor_match = None
        for entrada in filas_idx:
            ratio = _similitud(svc["alcance"], entrada["alcance"])
            if ratio > mejor_ratio:
                mejor_ratio = ratio
                mejor_match = entrada

        if mejor_match and mejor_ratio >= UMBRAL_SIMILITUD:
            svc["proveedor"] = mejor_match["proveedor"]
            svc["servicio"]  = mejor_match["servicio"]
            # Usa valor del Excel como fallback si el PDF no lo tenía
            if svc["valor_total_cop"] is None and mejor_match["valor"] is not None:
                svc["valor_total_cop"] = mejor_match["valor"]
                svc["valor_iva_cop"]   = round(mejor_match["valor"] * 0.19, 2)
                logger.info("Valor tomado del Excel para: %s", svc["alcance"][:60])
        else:
            logger.warning("Sin match Excel (ratio=%.2f) para: %s", mejor_ratio, svc["alcance"][:60])

    return servicios


# =============================================================
#  ENSAMBLADO FINAL
# =============================================================

def ensamblar_insumos(nombre_proyecto: str) -> dict:
    """
    Orquesta la extracción de todas las fuentes y construye
    el diccionario de insumos listo para el Módulo 2.
    """
    logger.info("=" * 55)
    logger.info("  EXTRACTOR DE INSUMOS UPME")
    logger.info("  Proyecto: %s", nombre_proyecto)
    logger.info("=" * 55)

    errores = []

    # ── FUENTE 1: Solenium JSON ──────────────────────────────
    sol = cargar_datos_solenium(nombre_proyecto)
    fpo_unergy         = sol.get("fpo_unergy")
    inicio_construccion = calcular_inicio_construccion(fpo_unergy)

    # NIT del primer inversionista (hardcodeado hasta que la API de dev lo exponga)
    nit_inversionista = "901938257"

    # ── FUENTE 2: PDF Plano de localizaciones ────────────────
    try:
        datos_plano = extraer_pdf_localizacion(PDF_LOCALIZACION)
    except Exception as e:
        logger.error("Error en PDF localización: %s", e)
        datos_plano = {}
        errores.append(f"pdf_localizacion: {e}")

    # ── FUENTE 2: PDF Informe eléctrico ─────────────────────
    try:
        datos_ele = extraer_pdf_informe_ele(PDF_INFORME_ELE)
    except Exception as e:
        logger.error("Error en PDF informe eléctrico: %s", e)
        datos_ele = {}
        errores.append(f"pdf_informe_ele: {e}")

    # ── CÁLCULOS DERIVADOS ───────────────────────────────────
    energia    = datos_ele.get("energia_generada_kwh_anio")
    capacidad  = datos_ele.get("capacidad_instalada_kwp")
    factor_planta = None
    if energia and capacidad and capacidad > 0:
        factor_planta = round((energia / (capacidad * 365 * 24)) * 100, 4)

    # ── DESCRIPCIÓN DEL PROYECTO ─────────────────────────────
    municipio    = sol.get("ciudad") or "—"
    departamento = sol.get("departamento") or "—"
    cap_str      = str(capacidad) if capacidad else "—"
    descripcion  = (
        f"El proyecto consiste en la instalación en suelo con seguidores solares "
        f"de un sistema interconectado de generación de energía fotovoltaica de "
        f"una capacidad instalada {cap_str} kW, que estará localizado en el "
        f"municipio de {municipio}, {departamento}, y estará interconectado a la "
        f"red de distribución eléctrica en media tensión de corriente alterna trifásica."
    )

    # ── FUENTE 3: Servicios ──────────────────────────────────
    try:
        servicios = extraer_servicios_pdf(DIR_SERVICIOS)
        servicios = enriquecer_servicios_excel(servicios, XLSX_BT)
    except Exception as e:
        logger.error("Error extrayendo servicios: %s", e)
        servicios = []
        errores.append(f"servicios: {e}")

    # ── DICCIONARIO FINAL ────────────────────────────────────
    insumos = {

        # --- Datos fijos (sin fuente) ---
        "tipo_proyecto_generacion_electrica": {
            "valor": "Generación Eléctrica",
            "fuente": None,
        },
        "etapa_del_proyecto": {
            "valor": "Inversión",
            "fuente": None,
        },
        "tipo_de_solicitante": {
            "valor": ["Principal", "Secundario"],
            "fuente": None,
        },
        "rol": {
            "valor": ["Dueño del Proyecto", "Entidad Bancaria", "Instalador",
                      "Importador", "Consultor", "Proveedor"],
            "fuente": None,
        },
        "nombre_contacto": {
            "valor": "Roberth Ricciulli",
            "fuente": None,
        },
        "telefono_celular_contacto": {
            "valor": "3054040085",
            "fuente": None,
        },
        "correo_electronico_contacto": {
            "valor": "roberth@solenium.co",
            "fuente": None,
        },
        "pais": {
            "valor": "Colombia",
            "fuente": None,
        },
        "tipo_de_generador": {
            "valor": "Generador",
            "fuente": None,
        },
        "tipo_fnce": {
            "valor": "Solar",
            "fuente": None,
        },
        "recurso_energetico": {
            "valor": "Sol",
            "fuente": None,
        },
        "tecnologia": {
            "valor": "Fotovoltaica",
            "fuente": None,
        },
        "sector": {
            "valor": "Generación Eléctrica",
            "fuente": None,
        },
        "zona": {
            "valor": "Sistema Interconectado Nacional (SIN)",
            "fuente": None,
        },
        "vida_util_proyecto_anios": {
            "valor": 25,
            "fuente": None,
        },
        "nivel_tension_kv": {
            "valor": "Nivel 2: Sistemas con tensión mayor o igual a 1 kV y menor a 30 kV",
            "fuente": None,
        },

        # --- Datos de Solenium JSON ---
        "nit_inversionista": {
            "valor": nit_inversionista,
            "fuente": "proyectos_solenium.json",
        },
        "departamento": {
            "valor": sol.get("departamento"),
            "fuente": "proyectos_solenium.json",
        },
        "municipio": {
            "valor": sol.get("ciudad"),
            "fuente": "proyectos_solenium.json",
        },
        "nombre_del_proyecto": {
            "valor": extraer_mgs_id(nombre_proyecto),
            "fuente": "proyectos_solenium.json",
        },
        "nombre_operador_de_red": {
            "valor": sol.get("operador_red_nombre"),
            "fuente": "proyectos_solenium.json",
        },
        "inicio_de_construccion": {
            "valor": inicio_construccion,
            "fuente": "proyectos_solenium.json",
        },
        "entrada_de_operacion": {
            "valor": fpo_unergy,
            "fuente": "proyectos_solenium.json",
        },

        # --- Datos del PDF plano de localizaciones ---
        "capacidad_instalada_kwp_plano": {
            "valor": datos_plano.get("capacidad_instalada_kwp_plano"),
            "fuente": "SDS4_Cope-CIV-PL-01_Plano de localizaciones y accesos.pdf",
        },
        "area_proyecto_m2": {
            "valor": datos_plano.get("area_proyecto_m2"),
            "fuente": "SDS4_Cope-CIV-PL-01_Plano de localizaciones y accesos.pdf",
        },

        # --- Datos del PDF informe eléctrico ---
        "energia_generada_kwh_anio": {
            "valor": datos_ele.get("energia_generada_kwh_anio"),
            "fuente": "Cope-INF-ELE-V2.pdf",
        },
        "eficiencia_planta_pct": {
            "valor": datos_ele.get("eficiencia_planta_pct"),
            "fuente": "Cope-INF-ELE-V2.pdf",
        },
        "capacidad_instalada_kwp": {
            "valor": capacidad,
            "fuente": "Cope-INF-ELE-V2.pdf",
        },

        # --- Datos calculados ---
        "factor_de_la_planta_pct": {
            "valor": factor_planta,
            "fuente": "Calculado: Energía / (Capacidad * 365 * 24)",
        },
        "descripcion_del_proyecto": {
            "valor": descripcion,
            "fuente": "Calculado con campos anteriores",
        },

        # --- Servicios BT ---
        "servicios_bt": {
            "valor": servicios,
            "fuente": "PDFs en Servicios Definitivos + Copey definitivo.xlsx",
        },

        # --- Metadatos ---
        "_proyecto": nombre_proyecto,
        "_errores":  errores,
    }

    # Resumen
    logger.info("=" * 55)
    logger.info("INSUMOS ENSAMBLADOS")
    logger.info("  Municipio        : %s", insumos["municipio"]["valor"])
    logger.info("  Departamento     : %s", insumos["departamento"]["valor"])
    logger.info("  NIT inversionista: %s", insumos["nit_inversionista"]["valor"])
    logger.info("  Capacidad (plano): %s kWp", insumos["capacidad_instalada_kwp_plano"]["valor"])
    logger.info("  Capacidad (ELE)  : %s kWp", insumos["capacidad_instalada_kwp"]["valor"])
    logger.info("  Energía P99      : %s kWh/año", insumos["energia_generada_kwh_anio"]["valor"])
    logger.info("  PR               : %s%%", insumos["eficiencia_planta_pct"]["valor"])
    logger.info("  Factor planta    : %s%%", insumos["factor_de_la_planta_pct"]["valor"])
    logger.info("  Inicio construc. : %s", insumos["inicio_de_construccion"]["valor"])
    logger.info("  Entrada operac.  : %s", insumos["entrada_de_operacion"]["valor"])
    logger.info("  Servicios BT     : %d alcances", len(servicios))
    if errores:
        logger.warning("  Errores          : %d", len(errores))
        for e in errores:
            logger.warning("    - %s", e)
    logger.info("=" * 55)

    return insumos


# =============================================================
#  PUNTO DE ENTRADA
# =============================================================
if __name__ == "__main__":
    insumos = ensamblar_insumos(NOMBRE_PROYECTO)

    salida = BASE_DIR / "data" / "insumos_MGS_0025.json"
    salida.parent.mkdir(parents=True, exist_ok=True)
    with open(salida, "w", encoding="utf-8") as f:
        json.dump(insumos, f, indent=2, ensure_ascii=False)

    logger.info("JSON guardado en: %s", salida)
